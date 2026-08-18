from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from jindu.engine.excel import export_xlsx
from jindu.engine.image import export_png
from jindu.engine.schedule import cascade_schedule, instantiate_template, load_templates


def test_templates_cover_trades():
    ids = {t["id"] for t in load_templates()}
    assert {"住宅楼", "市政道路", "装饰装修", "钢结构厂房"} <= ids


def test_residential_fs_and_rollup():
    project = instantiate_template(
        "住宅楼",
        name="试验楼",
        contract_start=date(2026, 1, 1),
        today=date(2026, 1, 20),
        demo_progress=False,
    )
    by_name = {t["name"]: t for t in project["tasks"]}
    assert by_name["施工准备"]["planned_start"] == "2026-01-01"
    assert by_name["施工准备"]["planned_end"] == "2026-01-15"
    assert by_name["支护结构施工"]["planned_start"] == "2026-01-16"
    pit = by_name["基坑支护与土方"]
    assert pit["summary"] is True
    assert pit["planned_start"] == by_name["支护结构施工"]["planned_start"]
    assert pit["planned_end"] == by_name["土方开挖与验槽"]["planned_end"]
    acc = by_name["竣工验收"]
    assert acc["planned_start"] > by_name["结构封顶"]["planned_end"]
    assert project["stats"]["task_count"] > 10
    assert any(t.get("critical") for t in project["tasks"] if not t.get("summary"))


def test_demo_progress_marks_started_work():
    project = instantiate_template(
        "住宅楼",
        name="进度样例",
        contract_start=date(2026, 1, 1),
        today=date(2026, 5, 1),
        demo_progress=True,
    )
    leaves = [t for t in project["tasks"] if not t.get("summary")]
    assert any(t["progress"] > 0 for t in leaves)
    assert project["stats"]["overall"] > 0


def test_cascade_respects_lag():
    tasks = [
        {
            "id": "a", "parent_id": "", "wbs": "1", "name": "A", "duration": 10,
            "offset": 0, "lag_days": 0, "predecessor_ids": [], "progress": 0,
        },
        {
            "id": "b", "parent_id": "", "wbs": "2", "name": "B", "duration": 5,
            "offset": 0, "lag_days": 2, "predecessor_ids": ["a"], "progress": 0,
        },
    ]
    out = {t["id"]: t for t in cascade_schedule(tasks, date(2026, 3, 1))}
    assert out["a"]["planned_end"] == "2026-03-10"
    assert out["b"]["planned_start"] == "2026-03-13"


def test_export_xlsx_and_png():
    project = instantiate_template(
        "市政道路",
        name="试验路",
        contract_start=date(2026, 2, 1),
        today=date(2026, 3, 1),
        demo_progress=True,
    )
    raw = export_xlsx(project, date(2026, 3, 1))
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) >= {"横道图", "任务表", "施工日志", "图例与说明"}
    assert "横道图" in (wb["横道图"]["A1"].value or "")
    png = export_png(project, date(2026, 3, 1))
    assert png[:8] == b"\x89PNG\r\n\x1a\n"
