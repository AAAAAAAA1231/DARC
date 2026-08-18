from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from zhiliang.engine.quality import enrich_project

FILL_HEADER = PatternFill("solid", fgColor="1B4F8A")
FILL_META = PatternFill("solid", fgColor="E8F0F8")
FILL_OPEN = PatternFill("solid", fgColor="DBEAFE")
FILL_LAG = PatternFill("solid", fgColor="FFEDD5")
FILL_MAJOR = PatternFill("solid", fgColor="FEE2E2")
FILL_DONE = PatternFill("solid", fgColor="DCFCE7")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
WHITE = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
FONT = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
TITLE = Font(name="微软雅黑", size=14, bold=True, color="1B4F8A")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fill_for(issue: dict[str, Any]) -> PatternFill | None:
    if issue.get("status") == "已闭合":
        return FILL_DONE
    if issue.get("severity") == "重大" and not issue.get("closed"):
        return FILL_MAJOR
    if issue.get("overdue"):
        return FILL_LAG
    if not issue.get("closed"):
        return FILL_OPEN
    return None


def _sheet_header(ws, title: str, meta: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = meta
    ws["A2"].font = FONT
    ws["A2"].fill = FILL_META


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for c, title in enumerate(headers, 1):
        cell = ws.cell(row, c, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = THIN


def export_xlsx(project: dict[str, Any], today: date | None = None) -> bytes:
    today = today or date.today()
    data = enrich_project(project, today)
    stats = data.get("stats") or {}
    meta = (
        f"工程：{data.get('name') or ''}　　地点：{data.get('location') or ''}　　"
        f"项目经理：{data.get('manager') or ''}　　质量员：{data.get('qc_lead') or ''}　　"
        f"导出日 {stats.get('today')}　　未闭合 {stats.get('open_count', 0)}　　超期 {stats.get('overdue_count', 0)}"
    )
    wb = Workbook()

    ledger = wb.active
    ledger.title = "问题台账"
    headers = [
        "编号", "问题", "专业", "部位", "等级", "来源", "发现日期", "整改期限",
        "责任人", "状态", "超期天数", "闭环步骤", "标准/规范", "实测/偏差",
        "原因(人机料法环)", "纠正措施", "预防措施", "复查结论",
    ]
    _sheet_header(ledger, f"{data.get('name') or '工程'} 质量问题台账", meta, len(headers))
    _write_header_row(ledger, 3, headers)
    for r, issue in enumerate(data.get("issues") or [], 4):
        cause = "；".join(
            f"{k}：{issue.get(f'cause_{name}')}"
            for k, name in (("人", "man"), ("机", "machine"), ("料", "material"), ("法", "method"), ("环", "env"))
            if issue.get(f"cause_{name}")
        )
        dev = " / ".join(x for x in (issue.get("actual"), issue.get("allowed"), issue.get("deviation")) if x)
        vals = [
            issue.get("no"), issue.get("title"), issue.get("specialty"), issue.get("location"),
            issue.get("severity"), issue.get("source"), issue.get("found_date"), issue.get("deadline"),
            issue.get("owner"), issue.get("status"),
            issue.get("overdue_days") or "", issue.get("loop_label"),
            issue.get("standard"), dev, cause,
            issue.get("corrective") or issue.get("rectify_plan"),
            issue.get("preventive"), issue.get("review_result"),
        ]
        fill = _fill_for(issue)
        for c, val in enumerate(vals, 1):
            cell = ledger.cell(r, c, val if val is not None else "")
            cell.font = FONT
            cell.alignment = LEFT if c in (2, 13, 14, 15, 16, 17, 18) else CENTER
            cell.border = THIN
            if fill:
                cell.fill = fill
    widths = [12, 22, 12, 18, 8, 12, 12, 12, 12, 10, 10, 10, 28, 18, 28, 28, 24, 18]
    for i, w in enumerate(widths, 1):
        ledger.column_dimensions[get_column_letter(i)].width = w
    ledger.freeze_panes = "A4"
    ledger.row_dimensions[3].height = 22

    notice = wb.create_sheet("整改通知单")
    _sheet_header(notice, f"{data.get('name') or '工程'} 质量整改通知单", meta + "　　（未闭合项）", 8)
    n_headers = ["编号", "部位", "问题", "等级", "期限", "责任人", "整改要求", "复查"]
    _write_header_row(notice, 3, n_headers)
    open_issues = [i for i in data.get("issues") or [] if not i.get("closed")]
    for r, issue in enumerate(open_issues, 4):
        vals = [
            issue.get("no"), issue.get("location"), issue.get("title"), issue.get("severity"),
            issue.get("deadline"), issue.get("owner"),
            issue.get("rectify_plan") or issue.get("corrective"),
            "待复查" if issue.get("status") == "待复查" else "未闭合",
        ]
        fill = _fill_for(issue)
        for c, val in enumerate(vals, 1):
            cell = notice.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT if c in (2, 3, 7) else CENTER
            cell.border = THIN
            if fill:
                cell.fill = fill
    for i, w in enumerate([12, 18, 22, 8, 12, 12, 40, 12], 1):
        notice.column_dimensions[get_column_letter(i)].width = w
    foot = 4 + max(len(open_issues), 1)
    notice.merge_cells(start_row=foot + 1, start_column=1, end_row=foot + 1, end_column=8)
    notice.cell(foot + 1, 1, "签发：质量员　　　　接收：责任班组　　　　监理：　　　　日期：").font = FONT

    corr = wb.create_sheet("纠偏记录")
    _sheet_header(corr, f"{data.get('name') or '工程'} 质量纠偏记录", "对照标准找偏差，定纠正/预防措施，复查闭合。", 10)
    c_headers = ["编号", "问题", "部位", "标准", "实测", "允许", "偏差", "原因分析", "纠正措施", "预防措施"]
    _write_header_row(corr, 3, c_headers)
    for r, issue in enumerate(data.get("issues") or [], 4):
        cause = "；".join(
            f"{k}{issue.get(f'cause_{name}')}"
            for k, name in (("人：", "man"), ("机：", "machine"), ("料：", "material"), ("法：", "method"), ("环：", "env"))
            if issue.get(f"cause_{name}")
        )
        vals = [
            issue.get("no"), issue.get("title"), issue.get("location"), issue.get("standard"),
            issue.get("actual"), issue.get("allowed"), issue.get("deviation"),
            cause, issue.get("corrective"), issue.get("preventive"),
        ]
        for c, val in enumerate(vals, 1):
            cell = corr.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = FILL_ALT
    for i, w in enumerate([12, 20, 16, 28, 16, 12, 12, 28, 28, 28], 1):
        corr.column_dimensions[get_column_letter(i)].width = w

    ins = wb.create_sheet("巡检记录")
    _sheet_header(ins, f"{data.get('name') or '工程'} 质量巡检记录", meta, 7)
    i_headers = ["日期", "类型", "区域", "检查人", "结论", "发现问题", "后续"]
    _write_header_row(ins, 3, i_headers)
    for r, rec in enumerate(data.get("inspections") or [], 4):
        vals = [rec.get("date"), rec.get("kind"), rec.get("area"), rec.get("inspector"), rec.get("result"), rec.get("findings"), rec.get("follow_up")]
        for c, val in enumerate(vals, 1):
            cell = ins.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT if c >= 6 else CENTER
            cell.border = THIN
    for i, w in enumerate([12, 12, 18, 12, 16, 40, 28], 1):
        ins.column_dimensions[get_column_letter(i)].width = w

    st = wb.create_sheet("统计")
    st["A1"] = "质量闭环统计"
    st["A1"].font = TITLE
    rows = [
        ("未闭合", stats.get("open_count")),
        ("超期", stats.get("overdue_count")),
        ("重大未闭合", stats.get("major_open")),
        ("本周新增", stats.get("new_week")),
        ("本周闭合", stats.get("closed_week")),
        ("闭合率%", stats.get("close_rate")),
        ("巡检次数", stats.get("inspect_count")),
    ]
    st["A3"] = "指标"
    st["B3"] = "数量"
    st["A3"].fill = FILL_HEADER
    st["B3"].fill = FILL_HEADER
    st["A3"].font = WHITE
    st["B3"].font = WHITE
    for i, (k, v) in enumerate(rows, 4):
        st.cell(i, 1, k).font = FONT
        st.cell(i, 2, v).font = FONT_BOLD
    st["A12"] = "按专业（问题条数）"
    st["A12"].font = FONT_BOLD
    r = 13
    for spec, n in (stats.get("by_specialty") or {}).items():
        st.cell(r, 1, spec).font = FONT
        st.cell(r, 2, n).font = FONT
        r += 1
    st.column_dimensions["A"].width = 18
    st.column_dimensions["B"].width = 12
    st["A22"] = "色块：蓝=未闭合　橙=超期　红=重大未闭合　绿=已闭合。本文件由「工程质量」本机软件导出。"
    st["A22"].font = FONT

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_name(project: dict[str, Any]) -> str:
    name = (project.get("name") or "工程质量").replace("/", "-").replace("\\", "-")
    return f"{name}-质量台账.xlsx"


def content_disposition(filename: str) -> str:
    return f"attachment; filename*=UTF-8''{quote(filename)}"
