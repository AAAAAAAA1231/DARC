from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from zhiliang.engine.excel import export_xlsx
from zhiliang.engine.quality import add_issue, demo_project, enrich_project, load_catalog, set_issue_status


def test_catalog_has_defects():
    cat = load_catalog()
    assert len(cat["defects"]) >= 8
    assert "混凝土工程" in cat["specialties"]


def test_demo_overdue_and_loop():
    today = date(2026, 8, 18)
    project = enrich_project(demo_project(today), today)
    stats = project["stats"]
    assert stats["open_count"] >= 1
    assert stats["overdue_count"] >= 1
    assert stats["issue_count"] == 5
    honey = next(i for i in project["issues"] if "蜂窝" in i["title"])
    assert honey["overdue"] is True
    assert honey["status"] == "待整改"
    closed = next(i for i in project["issues"] if i["status"] == "已闭合")
    assert closed["overdue"] is False


def test_status_close_loop():
    today = date(2026, 8, 18)
    project = demo_project(today)
    issue = add_issue(project, {"title": "试块漏做", "specialty": "混凝土工程", "found_date": "2026-08-10", "deadline": "2026-08-20"}, today)
    set_issue_status(issue, "整改中", {}, today)
    set_issue_status(issue, "待复查", {"rectify_desc": "已补做"}, today)
    set_issue_status(issue, "已闭合", {"reviewer": "质量员"}, today)
    assert issue["status"] == "已闭合"
    assert issue["rectify_done_date"]
    assert issue["review_date"] == "2026-08-18"


def test_export_xlsx_sheets():
    today = date(2026, 8, 18)
    raw = export_xlsx(demo_project(today), today)
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) >= {"问题台账", "整改通知单", "纠偏记录", "巡检记录", "统计"}
    assert "质量问题台账" in (wb["问题台账"]["A1"].value or "")
