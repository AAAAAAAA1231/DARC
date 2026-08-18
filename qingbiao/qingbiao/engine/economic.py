from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

HEADER_ALIASES = {
    "code": ("项目编码", "清单编码", "编码", "项目编号", "清单编号"),
    "name": ("项目名称", "名称", "项目特征", "清单名称", "分部分项"),
    "unit": ("计量单位", "单位", "单位名称"),
    "qty": ("工程量", "数量", "工程数量"),
    "unit_price": ("综合单价", "单价", "投标单价", "限价单价", "控制单价", "最高限价单价"),
    "amount": ("合价", "金额", "投标合价", "限价合价", "综合合价", "总价"),
}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v).strip())


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
    s = s.replace("元", "")
    try:
        return float(s)
    except ValueError:
        return None


def _map_headers(cells: list[Any]) -> dict[str, int]:
    texts = [_norm(c) for c in cells]
    mapping: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for i, t in enumerate(texts):
            if not t:
                continue
            if any(a in t or t == a for a in aliases):
                mapping[field_name] = i
                break
    return mapping


@dataclass
class BidItem:
    sheet: str
    row: int
    code: str
    name: str
    unit: str
    qty: float | None
    unit_price: float | None
    amount: float | None

    def key(self) -> str:
        if self.code:
            return self.code
        return f"{self.name}|{self.unit}"


@dataclass
class BidBook:
    bidder: str
    filename: str
    path: str = ""
    items: list[BidItem] = field(default_factory=list)
    sheets: list[str] = field(default_factory=list)

    def by_key(self) -> dict[str, BidItem]:
        out: dict[str, BidItem] = {}
        for it in self.items:
            k = it.key()
            if k and k not in out and it.unit_price is not None:
                out[k] = it
        return out

    def as_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["item_count"] = len(self.items)
        return d


def parse_excel_bid(path: Path, bidder: str) -> BidBook:
    wb = load_workbook(path, data_only=True, read_only=True)
    book = BidBook(bidder=bidder, filename=path.name, path=str(path), sheets=list(wb.sheetnames))
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header_map: dict[str, int] | None = None
        for ridx, row in enumerate(rows, start=1):
            cells = list(row)
            if header_map is None:
                if ridx > 20:
                    break
                mapping = _map_headers(cells)
                if "unit_price" in mapping and ("name" in mapping or "code" in mapping):
                    header_map = mapping
                continue
            name = _norm(cells[header_map["name"]]) if "name" in header_map and header_map["name"] < len(cells) else ""
            code = _norm(cells[header_map["code"]]) if "code" in header_map and header_map["code"] < len(cells) else ""
            if not name and not code:
                continue
            if name in HEADER_ALIASES["name"] or "合计" in name or "小计" in name:
                continue
            unit = _norm(cells[header_map["unit"]]) if "unit" in header_map and header_map["unit"] < len(cells) else ""
            qty = _num(cells[header_map["qty"]]) if "qty" in header_map and header_map["qty"] < len(cells) else None
            price = _num(cells[header_map["unit_price"]]) if header_map["unit_price"] < len(cells) else None
            amount = _num(cells[header_map["amount"]]) if "amount" in header_map and header_map["amount"] < len(cells) else None
            if price is None and amount is None:
                continue
            book.items.append(
                BidItem(
                    sheet=ws.title,
                    row=ridx,
                    code=code,
                    name=name or code,
                    unit=unit,
                    qty=qty,
                    unit_price=price,
                    amount=amount,
                )
            )
    wb.close()
    return book


def _near(a: float, b: float, pct: float, abs_tol: float) -> bool:
    if abs(a - b) <= abs_tol:
        return True
    base = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / base <= pct


def compare_to_limit(
    limit: BidBook,
    books: Iterable[BidBook],
    similar_pct: float = 0.005,
    abs_tol: float = 0.01,
) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    limit_map = limit.by_key()
    for book in books:
        if not book.items:
            findings.append(
                {
                    "module": "经济标",
                    "category": "未能解析清单",
                    "severity": "高",
                    "bidder": book.bidder,
                    "item_code": "",
                    "item_name": "",
                    "detail": "未识别到含「综合单价」的清单表。请确认是 Excel 工程量清单而不是封面或图片。",
                }
            )
            continue
        seen = set()
        for item in book.items:
            k = item.key()
            if not k or k in seen or item.unit_price is None:
                continue
            seen.add(k)
            lim = limit_map.get(k)
            if not lim or lim.unit_price is None:
                findings.append(
                    {
                        "module": "经济标",
                        "category": "限价漏项或编码不一致",
                        "severity": "中",
                        "bidder": book.bidder,
                        "item_code": item.code,
                        "item_name": item.name,
                        "detail": f"投标单价 {item.unit_price}，在最高投标限价中未找到对应清单项",
                    }
                )
                continue
            lp, tp = lim.unit_price, item.unit_price
            if abs(tp - lp) <= 1e-9:
                findings.append(
                    {
                        "module": "经济标",
                        "category": "与最高投标限价单价相同",
                        "severity": "高",
                        "bidder": book.bidder,
                        "item_code": item.code,
                        "item_name": item.name,
                        "detail": f"综合单价 {tp} 与限价 {lp} 完全相同（{item.sheet} 第{item.row}行）",
                    }
                )
            elif _near(tp, lp, similar_pct, abs_tol):
                pct = abs(tp - lp) / max(abs(lp), 1e-9) * 100
                findings.append(
                    {
                        "module": "经济标",
                        "category": "与最高投标限价单价相近",
                        "severity": "中",
                        "bidder": book.bidder,
                        "item_code": item.code,
                        "item_name": item.name,
                        "detail": f"综合单价 {tp} 与限价 {lp} 相差 {pct:.3f}%（{item.sheet} 第{item.row}行）",
                    }
                )
            if tp > lp + abs_tol:
                findings.append(
                    {
                        "module": "经济标",
                        "category": "超过最高投标限价",
                        "severity": "高",
                        "bidder": book.bidder,
                        "item_code": item.code,
                        "item_name": item.name,
                        "detail": f"综合单价 {tp} 高于限价 {lp}",
                    }
                )
        for k, lim in limit_map.items():
            if k not in seen:
                findings.append(
                    {
                        "module": "经济标",
                        "category": "投标漏项",
                        "severity": "中",
                        "bidder": book.bidder,
                        "item_code": lim.code,
                        "item_name": lim.name,
                        "detail": "最高投标限价有该项，投标文件未报出对应单价",
                    }
                )
    return findings


def cross_compare_prices(
    books: list[BidBook],
    similar_pct: float = 0.005,
    abs_tol: float = 0.01,
) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    keys: set[str] = set()
    maps = [(b, b.by_key()) for b in books]
    for _, m in maps:
        keys.update(m.keys())
    for k in sorted(keys):
        priced = [(b.bidder, m[k]) for b, m in maps if k in m and m[k].unit_price is not None]
        if len(priced) < 2:
            continue
        for i in range(len(priced)):
            for j in range(i + 1, len(priced)):
                ba, ia = priced[i]
                bb, ib = priced[j]
                pa, pb = ia.unit_price, ib.unit_price
                assert pa is not None and pb is not None
                if abs(pa - pb) <= 1e-9:
                    findings.append(
                        {
                            "module": "经济标",
                            "category": "多家投标单价相同",
                            "severity": "高",
                            "bidder": f"{ba} / {bb}",
                            "item_code": ia.code or ib.code,
                            "item_name": ia.name or ib.name,
                            "detail": f"综合单价均为 {pa}",
                        }
                    )
                elif _near(pa, pb, similar_pct, abs_tol):
                    pct = abs(pa - pb) / max(abs(pa), abs(pb), 1e-9) * 100
                    findings.append(
                        {
                            "module": "经济标",
                            "category": "多家投标单价相近",
                            "severity": "中",
                            "bidder": f"{ba} / {bb}",
                            "item_code": ia.code or ib.code,
                            "item_name": ia.name or ib.name,
                            "detail": f"{ba} 为 {pa}，{bb} 为 {pb}，相差 {pct:.3f}%",
                        }
                    )
    totals = []
    for b in books:
        amt = [it.amount for it in b.items if it.amount is not None]
        if amt:
            totals.append((b.bidder, round(sum(amt), 2)))
    for i in range(len(totals)):
        for j in range(i + 1, len(totals)):
            if totals[i][1] == totals[j][1]:
                findings.append(
                    {
                        "module": "经济标",
                        "category": "投标总价相同",
                        "severity": "高",
                        "bidder": f"{totals[i][0]} / {totals[j][0]}",
                        "item_code": "",
                        "item_name": "清单合价汇总",
                        "detail": f"解析到的合价合计均为 {totals[i][1]}",
                    }
                )
    return findings
