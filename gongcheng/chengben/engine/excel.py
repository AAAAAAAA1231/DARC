from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from chengben.engine.cost import enrich_project

FILL_HEADER = PatternFill("solid", fgColor="1B4F8A")
FILL_META = PatternFill("solid", fgColor="E8F0F8")
FILL_OVER = PatternFill("solid", fgColor="FEE2E2")
FILL_WARN = PatternFill("solid", fgColor="FFEDD5")
FILL_SAVE = PatternFill("solid", fgColor="DCFCE7")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
WHITE = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
FONT = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
TITLE = Font(name="微软雅黑", size=14, bold=True, color="1B4F8A")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _flag_fill(flag: str) -> PatternFill | None:
    return {"超支": FILL_OVER, "预警": FILL_WARN, "节约": FILL_SAVE}.get(flag)


def _header(ws, title: str, meta: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = meta
    ws["A2"].font = FONT
    ws["A2"].fill = FILL_META


def _heads(ws, row: int, headers: list[str]) -> None:
    for c, title in enumerate(headers, 1):
        cell = ws.cell(row, c, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = THIN


def _row(ws, r: int, vals: list[Any], fill: PatternFill | None = None) -> None:
    for c, val in enumerate(vals, 1):
        cell = ws.cell(r, c, val if val is not None else "")
        cell.font = FONT
        cell.alignment = RIGHT if isinstance(val, (int, float)) else LEFT
        cell.border = THIN
        if fill:
            cell.fill = fill
        elif r % 2 == 0:
            cell.fill = FILL_ALT


def _by_id(rows: list[dict[str, Any]], rid: str) -> dict[str, Any] | None:
    return next((x for x in rows if x.get("id") == rid), None)


def export_xlsx(project: dict[str, Any], today: date | None = None) -> bytes:
    today = today or date.today()
    data = enrich_project(project, today)
    stats = data.get("stats") or {}
    items = data.get("items") or []
    meta = (
        f"工程：{data.get('name') or ''}　　地点：{data.get('location') or ''}　　"
        f"目标 {stats.get('budget', 0):,.0f}　　动态 {stats.get('target', 0):,.0f}　　"
        f"已发生 {stats.get('actual', 0):,.0f}　　预计 {stats.get('forecast', 0):,.0f}　　"
        f"节超 {stats.get('deviation', 0):,.0f}　　导出 {stats.get('today')}"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "成本台账"
    headers = [
        "编码", "科目", "类别", "单位", "目标量", "目标单价", "目标金额", "变更", "动态成本",
        "已发生量", "已发生金额", "待发生", "预计总成本", "节超金额", "节超率", "状态", "责任人",
    ]
    _header(ws, f"{data.get('name') or '工程'} 成本台账", meta, len(headers))
    _heads(ws, 3, headers)
    for r, item in enumerate(items, 4):
        vals = [
            item.get("code"), item.get("name"), item.get("category"), item.get("unit"),
            item.get("budget_qty") or "", item.get("budget_price") or "", item.get("budget_amount"),
            item.get("change_amount"), item.get("target"),
            item.get("actual_qty") or "", item.get("actual_amount"), item.get("remain_amount"),
            item.get("forecast"), item.get("deviation"),
            f"{round((item.get('deviation_rate') or 0) * 100, 2)}%",
            item.get("flag"), item.get("owner") or "",
        ]
        _row(ws, r, vals, _flag_fill(str(item.get("flag") or "")))
    widths = [8, 16, 12, 8, 10, 10, 12, 10, 12, 10, 12, 12, 12, 12, 10, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    ana = wb.create_sheet("节超分析")
    _header(ana, f"{data.get('name') or '工程'} 节超分析", "正数=预计超支，负数=预计节约。量差/价差供纠偏参考。", 9)
    _heads(ana, 3, ["编码", "科目", "动态成本", "预计", "节超", "节超率", "量差", "价差", "纠偏重点"])
    for r, item in enumerate(items, 4):
        hint = ""
        if item.get("flag") in ("超支", "预警"):
            if abs(item.get("price_diff") or 0) >= abs(item.get("qty_diff") or 0) and item.get("budget_price"):
                hint = "偏价差，核采购价"
            else:
                hint = "偏量差或待发生偏高，核图纸/损耗"
        _row(
            ana,
            r,
            [
                item.get("code"), item.get("name"), item.get("target"), item.get("forecast"),
                item.get("deviation"), f"{round((item.get('deviation_rate') or 0) * 100, 2)}%",
                item.get("qty_diff") or "", item.get("price_diff") or "", hint,
            ],
            _flag_fill(str(item.get("flag") or "")),
        )
    for i, w in enumerate([8, 16, 12, 12, 12, 10, 10, 10, 28], 1):
        ana.column_dimensions[get_column_letter(i)].width = w

    corr = wb.create_sheet("纠偏记录")
    _header(corr, f"{data.get('name') or '工程'} 成本纠偏记录", meta, 10)
    _heads(corr, 3, ["编号", "日期", "科目", "类型", "纠偏金额", "原因", "措施", "责任人", "期限", "状态"])
    for r, row in enumerate(data.get("corrections") or [], 4):
        item = _by_id(items, row.get("item_id") or "")
        fill = FILL_WARN if row.get("overdue") else None
        _row(
            corr,
            r,
            [
                row.get("no"), row.get("date"), (item or {}).get("name"), row.get("kind"),
                row.get("deviation_amount"), row.get("cause"), row.get("action"),
                row.get("owner"), row.get("deadline"), row.get("status"),
            ],
            fill,
        )
    for i, w in enumerate([12, 12, 16, 10, 12, 28, 32, 10, 12, 10], 1):
        corr.column_dimensions[get_column_letter(i)].width = w

    logs = wb.create_sheet("发生明细")
    _header(logs, f"{data.get('name') or '工程'} 成本发生明细", meta, 7)
    _heads(logs, 3, ["日期", "科目", "类型", "数量", "金额", "单据号", "备注"])
    for r, row in enumerate(data.get("logs") or [], 4):
        item = _by_id(items, row.get("item_id") or "")
        _row(logs, r, [row.get("date"), (item or {}).get("name"), row.get("kind"), row.get("qty") or "", row.get("amount"), row.get("voucher"), row.get("notes")])
    for i, w in enumerate([12, 16, 12, 10, 12, 14, 32], 1):
        logs.column_dimensions[get_column_letter(i)].width = w

    chg = wb.create_sheet("签证变更")
    _header(chg, f"{data.get('name') or '工程'} 签证变更", "已批准的变更计入动态成本。", 7)
    _heads(chg, 3, ["编号", "日期", "内容", "金额", "对应科目", "已批准", "备注"])
    for r, row in enumerate(data.get("changes") or [], 4):
        item = _by_id(items, row.get("item_id") or "")
        _row(chg, r, [row.get("no"), row.get("date"), row.get("title"), row.get("amount"), (item or {}).get("name"), "是" if row.get("approved") else "否", row.get("notes")])
    for i, w in enumerate([12, 12, 24, 12, 16, 10, 32], 1):
        chg.column_dimensions[get_column_letter(i)].width = w

    st = wb.create_sheet("统计")
    st["A1"] = "成本纠偏统计"
    st["A1"].font = TITLE
    st["A3"] = "指标"
    st["B3"] = "金额/数量"
    st["A3"].fill = FILL_HEADER
    st["B3"].fill = FILL_HEADER
    st["A3"].font = WHITE
    st["B3"].font = WHITE
    rows = [
        ("目标成本", stats.get("budget")),
        ("变更合计", stats.get("change")),
        ("动态成本", stats.get("target")),
        ("已发生", stats.get("actual")),
        ("待发生", stats.get("remain")),
        ("预计总成本", stats.get("forecast")),
        ("节超（预计-动态）", stats.get("deviation")),
        ("节超率", f"{round((stats.get('deviation_rate') or 0) * 100, 2)}%"),
        ("预警科目", stats.get("warn_count")),
        ("超支科目", stats.get("over_count")),
        ("纠偏未闭合", stats.get("open_corr")),
        ("纠偏超期", stats.get("overdue_corr")),
    ]
    for i, (k, v) in enumerate(rows, 4):
        st.cell(i, 1, k).font = FONT
        st.cell(i, 2, v).font = FONT_BOLD
    st["A18"] = "按类别（动态 / 已发生 / 预计 / 节超）"
    st["A18"].font = FONT_BOLD
    r = 19
    for cat, bucket in (stats.get("by_category") or {}).items():
        st.cell(r, 1, cat).font = FONT
        st.cell(r, 2, bucket.get("target")).font = FONT
        st.cell(r, 3, bucket.get("actual")).font = FONT
        st.cell(r, 4, bucket.get("forecast")).font = FONT
        st.cell(r, 5, bucket.get("deviation")).font = FONT
        r += 1
    st["A32"] = "红=超支（≥10%）　橙=预警（5%～10%）　绿=节约（优于目标 2% 以上）。金额单位：元。"
    st["A32"].font = FONT
    st.column_dimensions["A"].width = 22
    for col in "BCDE":
        st.column_dimensions[col].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_name(project: dict[str, Any]) -> str:
    name = (project.get("name") or "工程成本").replace("/", "-").replace("\\", "-")
    return f"{name}-成本台账.xlsx"


def content_disposition(filename: str) -> str:
    return f"attachment; filename*=UTF-8''{quote(filename)}"
