from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from jindu.engine.schedule import enrich_project, inclusive_days, parse_date

FILL_HEADER = PatternFill("solid", fgColor="1B4F8A")
FILL_META = PatternFill("solid", fgColor="E8F0F8")
FILL_WEEKEND = PatternFill("solid", fgColor="F3F4F6")
FILL_TODAY = PatternFill("solid", fgColor="FDE68A")
FILL_PLAN = PatternFill("solid", fgColor="93C5FD")
FILL_PROG = PatternFill("solid", fgColor="1D4ED8")
FILL_DELAY = PatternFill("solid", fgColor="EA580C")
FILL_DONE = PatternFill("solid", fgColor="15803D")
FILL_CRIT = PatternFill("solid", fgColor="FEE2E2")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
WHITE_FONT = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
FONT = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_TITLE = Font(name="微软雅黑", size=14, bold=True, color="1B4F8A")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def _scale(start: date, end: date) -> str:
    span = inclusive_days(start, end)
    if span <= 90:
        return "day"
    return "week"


def _periods(start: date, end: date, scale: str) -> list[tuple[date, date, str]]:
    out: list[tuple[date, date, str]] = []
    cur = start
    while cur <= end:
        if scale == "day":
            label = f"{cur.month}/{cur.day}"
            out.append((cur, cur, label))
            cur += timedelta(days=1)
        else:
            stop = min(cur + timedelta(days=6), end)
            label = f"{cur.month}/{cur.day}"
            out.append((cur, stop, label))
            cur = stop + timedelta(days=1)
    return out


def _status_fill(status: str) -> PatternFill | None:
    if status == "滞后":
        return PatternFill("solid", fgColor="FFEDD5")
    if status == "已完成":
        return PatternFill("solid", fgColor="DCFCE7")
    if status == "进行中":
        return PatternFill("solid", fgColor="DBEAFE")
    return None


def _bar_fill(task: dict[str, Any], p0: date, p1: date, today: date) -> PatternFill | None:
    start = parse_date(task.get("planned_start"))
    end = parse_date(task.get("planned_end"))
    if not start or not end:
        return None
    if p1 < start or p0 > end:
        return None
    progress = int(task.get("progress") or 0)
    duration = max(inclusive_days(start, end), 1)
    done_end = start + timedelta(days=int(round((duration - 1) * progress / 100))) if progress else start - timedelta(days=1)
    status = task.get("status") or ""
    if status == "已完成" or progress >= 100:
        return FILL_DONE
    if progress > 0 and p0 <= done_end:
        return FILL_PROG
    if status == "滞后":
        return FILL_DELAY
    return FILL_PLAN


def export_xlsx(project: dict[str, Any], today: date | None = None) -> bytes:
    today = today or date.today()
    data = enrich_project(project, today)
    tasks = data.get("tasks") or []
    stats = data.get("stats") or {}
    starts = [parse_date(t.get("planned_start")) for t in tasks]
    ends = [parse_date(t.get("planned_end")) for t in tasks]
    dates = [d for d in starts + ends if d]
    if not dates:
        cs = parse_date(data.get("contract_start")) or today
        ce = parse_date(data.get("contract_end")) or today + timedelta(days=30)
        dates = [cs, ce]
    g0 = min(dates) - timedelta(days=2)
    g1 = max(dates) + timedelta(days=2)
    if today < g0:
        g0 = today - timedelta(days=1)
    if today > g1:
        g1 = today + timedelta(days=1)
    scale = _scale(g0, g1)
    periods = _periods(g0, g1, scale)

    wb = Workbook()
    ws = wb.active
    ws.title = "横道图"
    meta = [
        f"工程：{data.get('name') or ''}",
        f"地点：{data.get('location') or ''}",
        f"工期：{data.get('contract_start') or ''} ～ {data.get('contract_end') or ''}",
        f"总进度 {stats.get('overall', 0)}%　计划应完 {stats.get('planned_overall', 0)}%　SPI {stats.get('spi', 1)}　滞后 {stats.get('delayed_count', 0)} 项　导出日 {stats.get('today')}",
    ]
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{data.get('name') or '工程'} 施工进度横道图"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A2:G2")
    ws["A2"] = "　　".join(meta)
    ws["A2"].font = FONT
    ws["A2"].fill = FILL_META

    headers = ["WBS", "工作名称", "责任人", "计划开始", "计划完成", "工期", "进度%", "状态", "总时差", "前置"]
    fixed = len(headers)
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN
    for i, (_a, _b, label) in enumerate(periods):
        cell = ws.cell(3, fixed + 1 + i, label)
        cell.fill = FILL_HEADER
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        cell.border = THIN

    widths = [10, 22, 10, 12, 12, 8, 8, 10, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(fixed + 1 + i)].width = 2.6 if scale == "day" else 4.2

    by_id = {t["id"]: t for t in tasks}
    for r, task in enumerate(tasks, 4):
        pred_names = []
        for pid in task.get("predecessor_ids") or []:
            p = by_id.get(pid)
            if p:
                pred_names.append(p.get("wbs") or p.get("name") or "")
        values = [
            task.get("wbs") or "",
            ("　" if task.get("parent_id") else "") + (task.get("name") or ""),
            task.get("owner") or "",
            task.get("planned_start") or "",
            task.get("planned_end") or "",
            task.get("duration") or "",
            task.get("progress") or 0,
            task.get("status") or "",
            task.get("float_days") if task.get("float_days") is not None else "",
            "、".join(pred_names),
        ]
        row_fill = _status_fill(str(task.get("status") or ""))
        if not row_fill and r % 2 == 0:
            row_fill = FILL_ALT
        if task.get("critical") and not task.get("summary"):
            row_fill = FILL_CRIT
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.font = FONT_BOLD if task.get("summary") else FONT
            cell.alignment = LEFT if c == 2 else CENTER
            cell.border = THIN
            if row_fill:
                cell.fill = row_fill
        for i, (p0, p1, _label) in enumerate(periods):
            cell = ws.cell(r, fixed + 1 + i, "")
            cell.border = THIN
            fill = _bar_fill(task, p0, p1, today)
            if p0 <= today <= p1:
                cell.fill = FILL_TODAY if fill is None else fill
            elif fill is not None:
                cell.fill = fill
            elif p0.weekday() >= 5:
                cell.fill = FILL_WEEKEND

    ws.freeze_panes = "C4"
    ws.row_dimensions[3].height = 36
    ws.auto_filter.ref = f"A3:{get_column_letter(fixed)}{3 + max(len(tasks), 1)}"

    legend = wb.create_sheet("图例与说明")
    legend["A1"] = "横道图色块"
    legend["A1"].font = FONT_TITLE
    rows = [
        ("计划（未开始/未完成段）", "浅蓝"),
        ("已完成进度段", "深蓝 / 绿色"),
        ("滞后", "橙色"),
        ("关键线路任务", "浅红底"),
        ("当日列", "黄色"),
        ("周末", "浅灰"),
    ]
    fills = [FILL_PLAN, FILL_PROG, FILL_DELAY, FILL_CRIT, FILL_TODAY, FILL_WEEKEND]
    for i, ((name, _desc), fill) in enumerate(zip(rows, fills), 3):
        legend.cell(i, 1, name).font = FONT
        legend.cell(i, 2).fill = fill
        legend.cell(i, 2).border = THIN
    legend["A10"] = "本文件由「工程进度」本机软件导出，可在 Excel 中继续改色、打印。关键线路按完成-开始（FS）关系计算总时差。"
    legend["A10"].font = FONT
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 12

    log_ws = wb.create_sheet("施工日志")
    log_headers = ["日期", "天气", "气温", "人数", "完成工作", "存在问题", "明日计划", "记录人"]
    for c, title in enumerate(log_headers, 1):
        cell = log_ws.cell(1, c, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE_FONT
        cell.alignment = CENTER
    for r, log in enumerate(data.get("logs") or [], 2):
        vals = [log.get("date"), log.get("weather"), log.get("temperature"), log.get("manpower"), log.get("work"), log.get("issues"), log.get("tomorrow"), log.get("author")]
        for c, val in enumerate(vals, 1):
            cell = log_ws.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in enumerate([12, 8, 8, 8, 40, 28, 28, 10], 1):
        log_ws.column_dimensions[get_column_letter(col)].width = w

    task_ws = wb.create_sheet("任务表")
    cols = ["WBS", "工作名称", "责任人", "计划开始", "计划完成", "实际开始", "实际完成", "工期", "进度%", "状态", "滞后天数", "总时差", "关键", "备注"]
    for c, title in enumerate(cols, 1):
        cell = task_ws.cell(1, c, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE_FONT
    for r, task in enumerate(tasks, 2):
        vals = [
            task.get("wbs"), task.get("name"), task.get("owner"),
            task.get("planned_start"), task.get("planned_end"),
            task.get("actual_start"), task.get("actual_end"),
            task.get("duration"), task.get("progress"), task.get("status"),
            task.get("delayed_days"), task.get("float_days"),
            "是" if task.get("critical") else "", task.get("notes"),
        ]
        for c, val in enumerate(vals, 1):
            task_ws.cell(r, c, val if val is not None else "").font = FONT
    for col, w in enumerate([10, 22, 10, 12, 12, 12, 12, 8, 8, 10, 10, 8, 8, 24], 1):
        task_ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_name(project: dict[str, Any], ext: str) -> str:
    name = (project.get("name") or "工程进度").replace("/", "-").replace("\\", "-")
    return f"{name}-横道图.{ext}"


def content_disposition(filename: str) -> str:
    return f"attachment; filename*=UTF-8''{quote(filename)}"
