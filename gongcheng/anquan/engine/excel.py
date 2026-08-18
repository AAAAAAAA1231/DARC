from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from anquan.engine.safety import enrich_project

FILL_HEADER = PatternFill("solid", fgColor="9A3412")
FILL_META = PatternFill("solid", fgColor="FFEDD5")
FILL_OPEN = PatternFill("solid", fgColor="DBEAFE")
FILL_LAG = PatternFill("solid", fgColor="FDBA74")
FILL_MAJOR = PatternFill("solid", fgColor="FECACA")
FILL_DONE = PatternFill("solid", fgColor="DCFCE7")
FILL_ALT = PatternFill("solid", fgColor="FFF7ED")
WHITE = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
FONT = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
TITLE = Font(name="微软雅黑", size=14, bold=True, color="9A3412")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _fill_for(item: dict[str, Any]) -> PatternFill | None:
    if item.get("status") == "已闭合":
        return FILL_DONE
    if item.get("severity") == "重大隐患" and not item.get("closed"):
        return FILL_MAJOR
    if item.get("overdue"):
        return FILL_LAG
    if not item.get("closed"):
        return FILL_OPEN
    return None


def _sheet_header(ws, title: str, meta: str, cols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws["A2"] = meta
    ws["A2"].font = FONT
    ws["A2"].fill = FILL_META


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for c, title in enumerate(headers, 1):
        cell = ws.cell(row, c, title)
        cell.fill = FILL_HEADER
        cell.font = WHITE
        cell.alignment = CENTER
        cell.border = THIN


def export_xlsx(project: dict[str, Any], today: date | None = None) -> bytes:
    today = today or date.today()
    data = enrich_project(project, today)
    stats = data.get("stats") or {}
    meta = (
        f"工程：{data.get('name') or ''}　　地点：{data.get('location') or ''}　　"
        f"项目经理：{data.get('manager') or ''}　　安全员：{data.get('safety_lead') or ''}　　"
        f"导出日 {stats.get('today')}　　未闭合 {stats.get('open_count', 0)}　　"
        f"超期 {stats.get('overdue_count', 0)}　　重大 {stats.get('major_open', 0)}"
    )
    wb = Workbook()
    ledger = wb.active
    ledger.title = "隐患台账"
    headers = [
        "编号", "隐患", "类别", "部位", "等级", "来源", "发现日期", "整改期限",
        "责任人", "状态", "停工", "超期天数", "闭环", "依据", "现场情况",
        "原因(人机料法环)", "纠正措施", "预防措施", "复查结论",
    ]
    _sheet_header(ledger, f"{data.get('name') or '工程'} 安全隐患台账", meta, len(headers))
    _write_header_row(ledger, 3, headers)
    for r, item in enumerate(data.get("hazards") or [], 4):
        cause = "；".join(
            f"{k}：{item.get(f'cause_{name}')}"
            for k, name in (("人", "man"), ("机", "machine"), ("料", "material"), ("法", "method"), ("环", "env"))
            if item.get(f"cause_{name}")
        )
        vals = [
            item.get("no"), item.get("title"), item.get("category"), item.get("location"),
            item.get("severity"), item.get("source"), item.get("found_date"), item.get("deadline"),
            item.get("owner"), item.get("status"),
            "是" if item.get("stop_work") else "",
            item.get("overdue_days") or "", item.get("loop_label"),
            item.get("standard"), item.get("actual") or item.get("description"),
            cause, item.get("corrective") or item.get("rectify_plan"),
            item.get("preventive"), item.get("review_result"),
        ]
        fill = _fill_for(item)
        for c, val in enumerate(vals, 1):
            cell = ledger.cell(r, c, val if val is not None else "")
            cell.font = FONT
            cell.alignment = LEFT if c in (2, 14, 15, 16, 17, 18, 19) else CENTER
            cell.border = THIN
            if fill:
                cell.fill = fill
    widths = [12, 26, 12, 18, 10, 12, 12, 12, 12, 10, 8, 10, 8, 28, 22, 28, 28, 24, 18]
    for i, w in enumerate(widths, 1):
        ledger.column_dimensions[get_column_letter(i)].width = w
    ledger.freeze_panes = "A4"

    notice = wb.create_sheet("整改通知单")
    _sheet_header(notice, f"{data.get('name') or '工程'} 事故隐患整改通知单", meta + "　　（未闭合项）", 9)
    n_headers = ["编号", "部位", "隐患", "等级", "是否停工", "期限", "责任人", "整改要求", "复查"]
    _write_header_row(notice, 3, n_headers)
    open_items = [i for i in data.get("hazards") or [] if not i.get("closed")]
    for r, item in enumerate(open_items, 4):
        vals = [
            item.get("no"), item.get("location"), item.get("title"), item.get("severity"),
            "是" if item.get("stop_work") else "否", item.get("deadline"), item.get("owner"),
            item.get("rectify_plan") or item.get("corrective"),
            "待复查" if item.get("status") == "待复查" else "未闭合",
        ]
        fill = _fill_for(item)
        for c, val in enumerate(vals, 1):
            cell = notice.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT if c in (2, 3, 8) else CENTER
            cell.border = THIN
            if fill:
                cell.fill = fill
    for i, w in enumerate([12, 18, 24, 10, 10, 12, 12, 40, 12], 1):
        notice.column_dimensions[get_column_letter(i)].width = w
    foot = 4 + max(len(open_items), 1)
    notice.merge_cells(start_row=foot + 1, start_column=1, end_row=foot + 1, end_column=9)
    notice.cell(foot + 1, 1, "签发：安全员　　　　接收：责任班组　　　　监理：　　　　日期：").font = FONT

    corr = wb.create_sheet("纠偏记录")
    _sheet_header(corr, f"{data.get('name') or '工程'} 安全纠偏记录", "对照规范找偏差，定纠正/预防措施，复查闭合。重大隐患应立即停工整改。", 10)
    c_headers = ["编号", "隐患", "部位", "依据", "现场情况", "允许/要求", "偏差", "原因分析", "纠正措施", "预防措施"]
    _write_header_row(corr, 3, c_headers)
    for r, item in enumerate(data.get("hazards") or [], 4):
        cause = "；".join(
            f"{k}{item.get(f'cause_{name}')}"
            for k, name in (("人：", "man"), ("机：", "machine"), ("料：", "material"), ("法：", "method"), ("环：", "env"))
            if item.get(f"cause_{name}")
        )
        vals = [
            item.get("no"), item.get("title"), item.get("location"), item.get("standard"),
            item.get("actual") or item.get("description"), item.get("allowed"), item.get("deviation"),
            cause, item.get("corrective"), item.get("preventive"),
        ]
        for c, val in enumerate(vals, 1):
            cell = corr.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = FILL_ALT
    for i, w in enumerate([12, 22, 16, 28, 22, 14, 12, 28, 28, 28], 1):
        corr.column_dimensions[get_column_letter(i)].width = w

    ins = wb.create_sheet("安全巡查")
    _sheet_header(ins, f"{data.get('name') or '工程'} 安全巡查记录", meta, 7)
    i_headers = ["日期", "类型", "区域", "检查人", "结论", "发现问题", "后续"]
    _write_header_row(ins, 3, i_headers)
    for r, rec in enumerate(data.get("inspections") or [], 4):
        vals = [rec.get("date"), rec.get("kind"), rec.get("area"), rec.get("inspector"), rec.get("result"), rec.get("findings"), rec.get("follow_up")]
        for c, val in enumerate(vals, 1):
            cell = ins.cell(r, c, val or "")
            cell.font = FONT
            cell.alignment = LEFT if c >= 6 else CENTER
            cell.border = THIN
    for i, w in enumerate([12, 14, 20, 12, 16, 40, 28], 1):
        ins.column_dimensions[get_column_letter(i)].width = w

    st = wb.create_sheet("统计")
    st["A1"] = "安全闭环统计"
    st["A1"].font = TITLE
    st["A3"] = "指标"
    st["B3"] = "数量"
    st["A3"].fill = FILL_HEADER
    st["B3"].fill = FILL_HEADER
    st["A3"].font = WHITE
    st["B3"].font = WHITE
    rows = [
        ("未闭合", stats.get("open_count")),
        ("超期", stats.get("overdue_count")),
        ("重大隐患未闭合", stats.get("major_open")),
        ("涉及停工", stats.get("stop_count")),
        ("本周新增", stats.get("new_week")),
        ("本周闭合", stats.get("closed_week")),
        ("闭合率%", stats.get("close_rate")),
        ("巡查次数", stats.get("inspect_count")),
    ]
    for i, (k, v) in enumerate(rows, 4):
        st.cell(i, 1, k).font = FONT
        st.cell(i, 2, v).font = FONT_BOLD
    st["A13"] = "按类别（隐患条数）"
    st["A13"].font = FONT_BOLD
    r = 14
    for cat, n in (stats.get("by_category") or {}).items():
        st.cell(r, 1, cat).font = FONT
        st.cell(r, 2, n).font = FONT
        r += 1
    st.column_dimensions["A"].width = 20
    st.column_dimensions["B"].width = 12
    st["A28"] = "色块：蓝=未闭合　橙=超期　红=重大隐患　绿=已闭合。本文件由「工程安全」本机软件导出。"
    st["A28"].font = FONT

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_name(project: dict[str, Any]) -> str:
    name = (project.get("name") or "工程安全").replace("/", "-").replace("\\", "-")
    return f"{name}-隐患台账.xlsx"


def content_disposition(filename: str) -> str:
    return f"attachment; filename*=UTF-8''{quote(filename)}"
