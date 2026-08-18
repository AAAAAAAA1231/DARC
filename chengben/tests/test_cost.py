from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from chengben.engine.cost import (
    add_change,
    add_log,
    apply_demo_progress,
    enrich_project,
    instantiate_template,
    load_catalog,
)
from chengben.engine.excel import export_xlsx


def test_templates():
    ids = {t["id"] for t in load_catalog()["templates"]}
    assert {"住宅楼", "市政道路", "装饰装修"} <= ids


def test_budget_qty_times_price():
    project = instantiate_template("住宅楼", name="试")
    rebar = next(i for i in project["items"] if i["name"] == "钢筋")
    assert rebar["budget_amount"] == round(980 * 4100, 2)
    total = enrich_project(project, date(2026, 8, 18))["stats"]["budget"]
    assert total > 10_000_000


def test_demo_flags_overrun():
    today = date(2026, 8, 18)
    project = enrich_project(apply_demo_progress(instantiate_template("住宅楼", name="样例"), today), today)
    stats = project["stats"]
    assert stats["over_count"] >= 1
    assert stats["forecast"] > 0
    rebar = next(i for i in project["items"] if i["name"] == "钢筋")
    assert rebar["flag"] in ("超支", "预警")
    assert project["corrections"]


def test_log_and_change_roll_into_item():
    project = instantiate_template("市政道路", name="路")
    item = project["items"][0]
    before = item["actual_amount"]
    add_log(project, {"item_id": item["id"], "amount": 10000, "qty": 1, "kind": "进度计量"}, date(2026, 8, 1))
    assert item["actual_amount"] == before + 10000
    add_change(project, {"item_id": item["id"], "amount": 5000, "title": "增项", "approved": True}, date(2026, 8, 1))
    assert item["change_amount"] == 5000
    stats = enrich_project(project, date(2026, 8, 1))["stats"]
    assert stats["change"] == 5000


def test_xlsx_sheets():
    today = date(2026, 8, 18)
    raw = export_xlsx(apply_demo_progress(instantiate_template("装饰装修", name="精装"), today), today)
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) >= {"成本台账", "节超分析", "纠偏记录", "发生明细", "签证变更", "统计"}
    assert "成本台账" in (wb["成本台账"]["A1"].value or "")
