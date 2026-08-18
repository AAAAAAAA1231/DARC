from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from anquan.engine.excel import export_xlsx
from anquan.engine.safety import add_hazard, demo_project, enrich_project, load_catalog, set_status


def test_catalog_has_hazards():
    cat = load_catalog()
    assert len(cat["hazards"]) >= 10
    assert "高处作业" in cat["categories"]


def test_demo_major_and_overdue():
    today = date(2026, 8, 18)
    project = enrich_project(demo_project(today), today)
    stats = project["stats"]
    assert stats["hazard_count"] == 5
    assert stats["open_count"] >= 1
    assert stats["major_open"] >= 1
    assert stats["overdue_count"] >= 1
    edge = next(i for i in project["hazards"] if "临边" in i["title"] or "电梯井" in (i.get("location") or ""))
    assert edge["severity"] == "重大隐患"
    closed = next(i for i in project["hazards"] if i["status"] == "已闭合")
    assert closed["overdue"] is False


def test_close_loop():
    today = date(2026, 8, 18)
    project = demo_project(today)
    item = add_hazard(project, {"title": "通道堆砖", "category": "文明施工", "found_date": "2026-08-10", "deadline": "2026-08-20"}, today)
    set_status(item, "整改中", {}, today)
    set_status(item, "待复查", {"rectify_desc": "已清运"}, today)
    set_status(item, "已闭合", {"reviewer": "安全员"}, today)
    assert item["status"] == "已闭合"
    assert item["stop_work"] is False
    assert item["review_date"] == "2026-08-18"


def test_xlsx_sheets():
    today = date(2026, 8, 18)
    raw = export_xlsx(demo_project(today), today)
    wb = load_workbook(BytesIO(raw))
    assert set(wb.sheetnames) >= {"隐患台账", "整改通知单", "纠偏记录", "安全巡查", "统计"}
    assert "隐患台账" in (wb["隐患台账"]["A1"].value or "")
